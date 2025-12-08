python - <<EOF
from openpyxl import load_workbook
try:
    wb = load_workbook("drugs.xlsx")
    print("Sheets:", wb.sheetnames)
except Exception as e:
    print("ERROR:", e)
EOF
